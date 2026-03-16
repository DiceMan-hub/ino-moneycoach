#!/usr/bin/env python3
"""クレジットカード 乗継遅延・出航遅延 海外旅行・国内旅行自動付帯表をExcelに変換"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "クレカ遅延保険比較"

# ===== スタイル定義 =====
title_font = Font(name='Yu Gothic', size=12, bold=True)
header_font = Font(name='Yu Gothic', size=9, bold=True)
data_font = Font(name='Yu Gothic', size=8)
small_font = Font(name='Yu Gothic', size=7)
link_font = Font(name='Yu Gothic', size=6, color='0000FF')

green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
gray_fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
red_fill = PatternFill(start_color='FF6666', end_color='FF6666', fill_type='solid')
orange_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
light_blue_fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
legend_green = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')
legend_gray = PatternFill(start_color='A6A6A6', end_color='A6A6A6', fill_type='solid')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)


def apply_style(cell, font=None, fill=None, border=None, alignment=None):
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if alignment:
        cell.alignment = alignment


def set_cell(ws, row, col, value, font=data_font, fill=None, border=thin_border, alignment=center_align):
    cell = ws.cell(row=row, column=col, value=value)
    apply_style(cell, font=font, fill=fill, border=border, alignment=alignment)
    return cell


def merge_and_set(ws, row1, col1, row2, col2, value, font=header_font, fill=None, border=thin_border, alignment=center_align):
    if row1 != row2 or col1 != col2:
        ws.merge_cells(start_row=row1, start_column=col1, end_row=row2, end_column=col2)
    cell = ws.cell(row=row1, column=col1, value=value)
    apply_style(cell, font=font, fill=fill, border=border, alignment=alignment)
    # Apply border to all cells in merged range
    for r in range(row1, row2 + 1):
        for c in range(col1, col2 + 1):
            ws.cell(row=r, column=c).border = border
            if fill:
                ws.cell(row=r, column=c).fill = fill
    return cell


# ===== タイトル行 =====
row = 1
merge_and_set(ws, 1, 1, 1, 20, 'クレジットカード　乗継遅延・出航遅延　海外旅行・国内旅行自動付帯表',
              font=title_font, fill=None, border=None, alignment=left_align)

# 凡例
set_cell(ws, 1, 21, '緑は自動付帯', font=Font(name='Yu Gothic', size=8, bold=True, color='FFFFFF'),
         fill=legend_green, alignment=center_align)
set_cell(ws, 1, 22, '灰は利用付帯', font=Font(name='Yu Gothic', size=8, bold=True),
         fill=legend_gray, alignment=center_align)

# ===== 行ラベル定義 =====
ROW_LABELS_COL_A = ['', '乗継遅延/出航遅延', '', '', '', '', '', '携行品損害']
ROW_LABELS_COL_B = ['ホテル', '交通費、キャンセル料', '食事代', '受託遅延', '出航遅延',
                     '受託手荷物遅延', '受託手荷物紛失', '品用/偶然・損害']

# ===================================================================
# セクション1
# ===================================================================
sec1_start = 3
r = sec1_start

# ----- カード1: J-WESTゴールドカード VISA/MASTER -----
# ----- カード2: J-WESTゴールドカード JCB -----
# ----- カード3: auPAYゴールドカード -----
# ----- カード4: ★HeartOneカード/プラチナカード -----
# ----- カード5: ★dカードゴールド/プラチナ -----

# セクション1のカード定義
sec1_cards = [
    {
        'name': 'J-WESTゴールドカード',
        'brand': 'VISA/MASTER',
        'fee': 'ゴールド11,000',
        'fee2': 'エクスプレス12,100',
        'fee_fill': None,
        'fee2_fill': yellow_fill,
    },
    {
        'name': 'J-WESTゴールドカード',
        'brand': 'JCB',
        'fee': 'ゴールド11,000',
        'fee2': 'エクスプレス12,100',
        'fee_fill': None,
        'fee2_fill': yellow_fill,
    },
    {
        'name': 'auPAYゴールドカード',
        'brand': 'VISA/MASTER/AMEX',
        'fee': '11,000',
        'fee2': '',
        'fee_fill': None,
        'fee2_fill': None,
    },
    {
        'name': '★HeartOneカード/プラチナカード',
        'brand': 'AMEX',
        'fee': 'ノーマル3,300',
        'fee2': 'プラチナ22,000',
        'fee_fill': orange_fill,
        'fee2_fill': red_fill,
    },
    {
        'name': '★dカードゴールド/プラチナ',
        'brand': 'VISA/MASTER',
        'fee': '11,000',
        'fee2': '',
        'fee_fill': None,
        'fee2_fill': None,
    },
]

# セクション1のデータ (8行 × カード数×4列)
# 各カードは [海外本人, 海外家族, 国内本人, 国内家族]
# 色: 'g'=green(自動付帯), 'r'=gray(利用付帯), ''=なし
sec1_data = [
    # Row 0: ホテル
    # J-WEST VM      J-WEST JCB      auPAY           HeartOne                    dカード
    [('30K','g'),  ('-',''),  ('-',''),  ('-',''),    ('-',''),  ('-',''),  ('-',''),  ('-',''),    ('-',''),  ('-',''),  ('-',''),  ('-',''),    ('-',''),  ('20K','g'),  ('20K','g'),  ('-',''),    ('30K','g'),  ('30K','g'),  ('20K','r'),  ('20K','r')],
    # Row 1: 交通費、キャンセル料
    [('10K','g'),  ('-',''),  ('-',''),  ('-',''),    ('-',''),  ('-',''),  ('-',''),  ('-',''),    ('-',''),  ('-',''),  ('-',''),  ('-',''),    ('-',''),  ('20K','g'),  ('20K','g'),  ('-',''),    ('10K','g'),  ('10K','g'),  ('-',''),  ('-','')],
    # Row 2: 食事代
    [('10K','g'),  ('-',''),  ('-',''),  ('-',''),    ('-',''),  ('-',''),  ('-',''),  ('-',''),    ('-',''),  ('-',''),  ('-',''),  ('-',''),    ('-',''),  ('万/食20K','g'),  ('万/食20K','g'),  ('-',''),    ('5K','g'),  ('5K','g'),  ('10K','r'),  ('10K','r')],
    # Row 3: 受託遅延
    [('-',''),  ('-',''),  ('20K','r'),  ('-',''),    ('-',''),  ('-',''),  ('万/食20K','g'),  ('-',''),    ('-',''),  ('-',''),  ('万/食20K','g'),  ('万/食20K','g'),    ('30K','g'),  ('30K','g'),  ('-',''),  ('-',''),    ('-',''),  ('-',''),  ('20K','r'),  ('20K','r')],
    # Row 4: 出航遅延
    [('-',''),  ('-',''),  ('10K','r'),  ('-',''),    ('-',''),  ('-',''),  ('食20K','g'),  ('-',''),    ('-',''),  ('-',''),  ('食20K','g'),  ('食20K','g'),    ('30K','g'),  ('30K','g'),  ('-',''),  ('-',''),    ('-',''),  ('-',''),  ('-',''),  ('-','')],
    # Row 5: 受託手荷物遅延
    [('30K','g'),  ('-',''),  ('10K','r'),  ('-',''),    ('-',''),  ('-',''),  ('20K','g'),  ('-',''),    ('-',''),  ('-',''),  ('30K','g'),  ('-',''),    ('100K','g'),  ('100K','g'),  ('-',''),  ('-',''),    ('-',''),  ('-',''),  ('20K','r'),  ('20K','r')],
    # Row 6: 受託手荷物紛失
    [('-',''),  ('-',''),  ('-',''),  ('-',''),    ('-',''),  ('-',''),  ('-',''),  ('-',''),    ('-',''),  ('-',''),  ('-',''),  ('-',''),    ('100K×3','g'),  ('100K×3','g'),  ('-',''),  ('-',''),    ('-',''),  ('-',''),  ('-',''),  ('-','')],
    # Row 7: 携行品損害
    [('500K','g'),  ('-',''),  ('-',''),  ('-',''),    ('500K','g'),  ('500K','g'),  ('-',''),  ('-',''),    ('500K','g'),  ('500K','g'),  ('-',''),  ('-',''),    ('@300K/500K','g'),  ('150K','g'),  ('-',''),  ('-',''),    ('@300K/300K','g'),  ('150K','g'),  ('-',''),  ('-','')],
]

sec1_links = [
    'https://www.wester.jr-odekake.net/point/insurance/',
    'https://www.wester.jr-odekake.net/point/insurance/',
    'https://www.kddi-fs.com/...',
    'https://www.daiwa...',
    'https://dcard.docomo.ne.jp/st/...',
]


def write_section(ws, start_row, cards, data, links, label_col_start=1):
    """セクションを書き込む"""
    r = start_row
    num_cards = len(cards)
    data_col_start = label_col_start + 2  # ラベル列2つ分

    # ----- カード名行 -----
    col = data_col_start
    for card in cards:
        merge_and_set(ws, r, col, r, col + 3, card['name'],
                      font=header_font, fill=light_blue_fill)
        col += 4
    r += 1

    # ----- ブランド行 -----
    col = data_col_start
    for card in cards:
        merge_and_set(ws, r, col, r, col + 3, card['brand'],
                      font=data_font, fill=None)
        col += 4
    r += 1

    # ----- 年会費行 -----
    set_cell(ws, r, label_col_start, '', font=header_font)
    set_cell(ws, r, label_col_start + 1, '年会費', font=header_font)
    col = data_col_start
    for card in cards:
        if card.get('fee2'):
            merge_and_set(ws, r, col, r, col + 1, card['fee'],
                          font=data_font, fill=card.get('fee_fill'))
            merge_and_set(ws, r, col + 2, r, col + 3, card['fee2'],
                          font=data_font, fill=card.get('fee2_fill'))
        else:
            merge_and_set(ws, r, col, r, col + 3, card['fee'],
                          font=data_font, fill=card.get('fee_fill'))
        col += 4
    r += 1

    # ----- 補償内容行 -----
    set_cell(ws, r, label_col_start, '', font=header_font)
    set_cell(ws, r, label_col_start + 1, '補償内容', font=header_font)
    col = data_col_start
    for card in cards:
        merge_and_set(ws, r, col, r, col + 1, '海外', font=header_font)
        merge_and_set(ws, r, col + 2, r, col + 3, '国内', font=header_font)
        col += 4
    r += 1

    # ----- 対象行 -----
    set_cell(ws, r, label_col_start, '', font=header_font)
    set_cell(ws, r, label_col_start + 1, '対象', font=header_font)
    col = data_col_start
    for card in cards:
        set_cell(ws, r, col, '本人', font=data_font)
        set_cell(ws, r, col + 1, '家族', font=data_font)
        set_cell(ws, r, col + 2, '本人', font=data_font)
        set_cell(ws, r, col + 3, '家族', font=data_font)
        col += 4
    r += 1

    # ----- データ行 (8行) -----
    for i in range(8):
        # ラベル列A (カテゴリ)
        set_cell(ws, r, label_col_start, ROW_LABELS_COL_A[i], font=data_font, alignment=left_align)
        # ラベル列B (サブ項目)
        set_cell(ws, r, label_col_start + 1, ROW_LABELS_COL_B[i], font=data_font, alignment=left_align)

        # データ列
        col = data_col_start
        for j in range(len(data[i])):
            val, color = data[i][j]
            fill = None
            if color == 'g':
                fill = green_fill
            elif color == 'r':
                fill = gray_fill
            set_cell(ws, r, col, val, font=data_font, fill=fill)
            col += 1
        r += 1

    # ----- リンク行 -----
    set_cell(ws, r, label_col_start, '', font=data_font)
    set_cell(ws, r, label_col_start + 1, 'リンク', font=data_font)
    col = data_col_start
    for idx, link in enumerate(links):
        merge_and_set(ws, r, col, r, col + 3, link,
                      font=link_font, alignment=Alignment(horizontal='left', vertical='center', wrap_text=True))
        col += 4
    r += 1

    return r


# セクション1書き込み
sec1_end = write_section(ws, sec1_start, sec1_cards, sec1_data, sec1_links)

# ===================================================================
# セクション2
# ===================================================================
sec2_start = sec1_end + 2

sec2_cards = [
    {
        'name': 'AMEXプラチナカード/ANA AMEXプレミア',
        'brand': 'AMEX',
        'fee': '143,000',
        'fee2': '',
        'fee_fill': red_fill,
        'fee2_fill': None,
    },
    {
        'name': 'エポスプラチナカード',
        'brand': 'VISA',
        'fee': '20,000',
        'fee2': '',
        'fee_fill': None,
        'fee2_fill': None,
    },
    {
        'name': 'エポスゴールドカード',
        'brand': 'VISA',
        'fee': '',
        'fee2': '',
        'fee_fill': None,
        'fee2_fill': None,
    },
    {
        'name': 'オリコプラチナカード',
        'brand': 'MASTER',
        'fee': '20,370',
        'fee2': '',
        'fee_fill': None,
        'fee2_fill': None,
    },
    {
        'name': '★JAL/ANAゴールド/プラチナ',
        'brand': 'JCB',
        'fee': 'ゴールド17,600',
        'fee2': 'プラチナ34,100',
        'fee_fill': yellow_fill,
        'fee2_fill': red_fill,
    },
    {
        'name': '★CLUB-Aゴールド/プラチナ JCBのみ',
        'brand': 'JCB',
        'fee': 'ゴールド17,600',
        'fee2': 'プラチナ34,100',
        'fee_fill': yellow_fill,
        'fee2_fill': red_fill,
    },
    {
        'name': 'JCB/ダイヤモンド',
        'brand': 'JCB',
        'fee': 'プラチナ/海外',
        'fee2': '',
        'fee_fill': None,
        'fee2_fill': None,
    },
]

sec2_data = [
    # Row 0: ホテル
    [('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-','')],
    # Row 1: 交通費、キャンセル料
    [('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('24K','g'), ('-',''), ('30K','r'), ('-','')],
    # Row 2: 食事代
    [('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('24K','g'), ('-',''), ('30K','r'), ('-','')],
    # Row 3: 受託遅延
    [('万/食30K','g'), ('-',''), ('-',''), ('-',''),    ('万/食30K','g'), ('-',''), ('万/食30K','g'), ('-',''),    ('万/食30K','g'), ('-',''), ('万/食30K','g'), ('-',''),    ('万/食30K','g'), ('-',''), ('万/食30K','g'), ('-',''),    ('万/食20K','g'), ('万/食20K','g'), ('-',''), ('-',''),    ('-',''), ('万/食20K','r'), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-','')],
    # Row 4: 出航遅延
    [('食30K','g'), ('-',''), ('-',''), ('-',''),    ('20K','g'), ('20K','g'), ('-',''), ('-',''),    ('食30K','g'), ('-',''), ('食30K','g'), ('-',''),    ('食30K','g'), ('-',''), ('食30K','g'), ('-',''),    ('食20K','g'), ('食20K','g'), ('-',''), ('-',''),    ('-',''), ('食20K','r'), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-','')],
    # Row 5: 受託手荷物遅延
    [('30K','g'), ('-',''), ('-',''), ('-',''),    ('100K','g'), ('100K','g'), ('-',''), ('-',''),    ('30K','g'), ('-',''), ('30K','g'), ('-',''),    ('30K','g'), ('-',''), ('30K','g'), ('-',''),    ('20K','g'), ('20K','g'), ('-',''), ('-',''),    ('-',''), ('20K','r'), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-','')],
    # Row 6: 受託手荷物紛失
    [('60K','g'), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('40K','g'), ('40K','g'), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('40K','r'), ('-','')],
    # Row 7: 携行品損害
    [('1000K','g'), ('1000K','g'), ('-',''), ('-',''),    ('1000K','g'), ('1000K','g'), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('100K×10','g'), ('100K×10','g'), ('-',''), ('-',''),    ('500K','g'), ('-',''), ('-',''), ('-',''),    ('-',''), ('500K','r'), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-','')],
]

sec2_links = [
    'https://www.americanexpress.com/...',
    'https://www.eposcard.co.jp/...',
    '',
    'https://www.orico.tv/...',
    'https://hoken.jcb.co.jp/...',
    'https://www.jal.co.jp/...',
    '',
]

sec2_end = write_section(ws, sec2_start, sec2_cards, sec2_data, sec2_links)

# ===================================================================
# セクション3
# ===================================================================
sec3_start = sec2_end + 2

sec3_cards = [
    {
        'name': '三菱UFJニコスゴールドカード',
        'brand': 'VISA/MASTER/JCB',
        'fee': '11,000',
        'fee2': '',
        'fee_fill': None,
        'fee2_fill': None,
    },
    {
        'name': '三菱UFJニコスプラチナカード',
        'brand': 'AMEX',
        'fee': '22,000',
        'fee2': '',
        'fee_fill': yellow_fill,
        'fee2_fill': None,
    },
    {
        'name': '三菱UFJニコスゴールドカード',
        'brand': 'AMEX',
        'fee': '11,000',
        'fee2': '',
        'fee_fill': None,
        'fee2_fill': None,
    },
    {
        'name': '★出光カードゴールド/プラチナ',
        'brand': 'VISA/AMEX',
        'fee': 'ゴールド11,000',
        'fee2': 'プラチナ22,000',
        'fee_fill': orange_fill,
        'fee2_fill': red_fill,
    },
]

sec3_data = [
    # Row 0: ホテル
    [('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('30,000','g'), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-','')],
    # Row 1: 交通費、キャンセル料
    [('30K','g'), ('-',''), ('-',''), ('-',''),    ('-',''), ('10K','r'), ('-',''), ('-',''),    ('10K','g'), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('20K','r'), ('-','')],
    # Row 2: 食事代
    [('10K','g'), ('-',''), ('-',''), ('-',''),    ('-',''), ('10K','r'), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-','')],
    # Row 3: 受託遅延
    [('8K','g'), ('-',''), ('-',''), ('-',''),    ('20K','g'), ('30K','r'), ('-',''), ('-',''),    ('30K','g'), ('-',''), ('20K','r'), ('-',''),    ('20K','g'), ('20K','g'), ('20K','r'), ('-','')],
    # Row 4: 出航遅延
    [('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('10K','r'), ('-',''), ('-',''),    ('-',''), ('-',''), ('10K','r'), ('-',''),    ('20K','g'), ('20K','g'), ('20K','r'), ('-','')],
    # Row 5: 受託手荷物遅延
    [('30K','g'), ('-',''), ('-',''), ('-',''),    ('30K','g'), ('15K','r'), ('-',''), ('-',''),    ('30K','g'), ('-',''), ('10K','r'), ('-',''),    ('-',''), ('-',''), ('100K','r'), ('-','')],
    # Row 6: 受託手荷物紛失
    [('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-',''),    ('-',''), ('-',''), ('-',''), ('-','')],
    # Row 7: 携行品損害
    [('500K','g'), ('-',''), ('-',''), ('-',''),    ('500K','g'), ('-',''), ('-',''), ('-',''),    ('500K','g'), ('-',''), ('-',''), ('-',''),    ('500K','g'), ('-',''), ('-',''), ('-','')],
]

sec3_links = [
    'https://www.cr.mufg.jp/mycard/service/other/insurance/delay/index.html',
    'https://www.cr.mufg.jp/mycard/service/other/insurance/delay/index.html',
    'https://www.cr.mufg.jp/mycard/service/other/insurance/delay/index.html',
    'https://www.idemitsucard.com/privilege/hou_in_s/insure/',
]

sec3_end = write_section(ws, sec3_start, sec3_cards, sec3_data, sec3_links)

# ===== 列幅設定 =====
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 16
for i in range(3, 32):
    ws.column_dimensions[get_column_letter(i)].width = 10

# ===== 行の高さ設定 =====
for r in range(1, sec3_end + 1):
    ws.row_dimensions[r].height = 18

# 保存
output_path = '/home/user/ino-moneycoach/クレジットカード_乗継遅延_出航遅延_保険比較表.xlsx'
wb.save(output_path)
print(f"Excel file saved to: {output_path}")
