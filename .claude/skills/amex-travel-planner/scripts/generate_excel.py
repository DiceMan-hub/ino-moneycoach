#!/usr/bin/env python3
"""
Amex Travel Planner - Excel Generator
FHR/THC旅行プランニング用エクセルファイル生成スクリプト

Usage:
    python generate_excel.py --data <json_data_file> --output <output_path>
    
    または直接JSONデータを渡す:
    python generate_excel.py --json '{"fhr": [...], "thc": [...], "plan": [...]}'
"""

import json
import sys
import argparse
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        GradientFill
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.styles.numbers import FORMAT_NUMBER_COMMA_SEPARATED1
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError:
    print("openpyxl が必要です: pip install openpyxl --break-system-packages")
    sys.exit(1)


# ============================
# カラー定義
# ============================
COLORS = {
    # 背景色
    "fhr_bg":       "FFF8DC",   # シャンパンゴールド (FHR)
    "thc_bg":       "DCE9FF",   # ライトブルー (THC)
    "plan_bg":      "DCFFDC",   # ライトグリーン (推奨プラン)
    "offer_bg":     "FFE8CC",   # ライトオレンジ (特別オファーあり)
    "header_bg":    "1F3864",   # ダークネイビー (ヘッダー)
    "sub_header":   "2E74B5",   # ブルー (サブヘッダー)
    "marriott":     "CC0000",   # マリオット赤
    "ihg":          "006DB7",   # IHGブルー
    "hyatt":        "003865",   # ハイアットネイビー
    "hilton":       "002B7F",   # ヒルトンブルー
    "highlight":    "FFD700",   # ゴールド (最安値ハイライト)
    "white":        "FFFFFF",
    "light_gray":   "F5F5F5",
    "border_color": "BFBFBF",
}

CHAIN_PRIORITY = {
    "marriott": 1, "marriott bonvoy": 1, "marriott/sph": 1,
    "ihg": 2, "intercontinental": 2, "kimpton": 2, "crowne plaza": 2,
    "hyatt": 3, "park hyatt": 3, "grand hyatt": 3, "andaz": 3,
    "hilton": 4, "waldorf": 4, "waldorf astoria": 4, "conrad": 4,
}


def get_chain_priority(chain_name):
    """チェーン優先度を返す (1=最高, 5=その他)"""
    if not chain_name:
        return 5
    chain_lower = chain_name.lower()
    for key, priority in CHAIN_PRIORITY.items():
        if key in chain_lower:
            return priority
    return 5


def get_chain_color(chain_name):
    """チェーンに応じた文字色を返す"""
    if not chain_name:
        return "000000"
    chain_lower = chain_name.lower()
    if "marriott" in chain_lower or "sheraton" in chain_lower or "westin" in chain_lower:
        return COLORS["marriott"]
    elif "ihg" in chain_lower or "intercontinental" in chain_lower or "kimpton" in chain_lower:
        return COLORS["ihg"]
    elif "hyatt" in chain_lower or "park hyatt" in chain_lower or "andaz" in chain_lower:
        return COLORS["hyatt"]
    elif "hilton" in chain_lower or "waldorf" in chain_lower or "conrad" in chain_lower:
        return COLORS["hilton"]
    return "000000"


def apply_header_style(cell, text=None):
    """ヘッダーセルにスタイルを適用"""
    if text is not None:
        cell.value = text
    cell.font = Font(name="Arial", bold=True, color=COLORS["white"], size=11)
    cell.fill = PatternFill(start_color=COLORS["header_bg"],
                            end_color=COLORS["header_bg"], fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        bottom=Side(style="medium", color=COLORS["border_color"]),
        right=Side(style="thin", color=COLORS["border_color"])
    )


def apply_data_style(cell, bg_color=None, bold=False, font_color="000000",
                     number_format=None, align="left"):
    """データセルにスタイルを適用"""
    cell.font = Font(name="Arial", bold=bold, color=font_color, size=10)
    if bg_color:
        cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = Border(
        bottom=Side(style="thin", color=COLORS["border_color"]),
        right=Side(style="thin", color=COLORS["border_color"])
    )
    if number_format:
        cell.number_format = number_format


def auto_fit_columns(ws, min_width=10, max_width=40):
    """列幅を自動調整"""
    for column_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value:
                # 日本語文字を考慮（2バイト文字は2倍の幅）
                text = str(cell.value)
                length = sum(2 if ord(c) > 127 else 1 for c in text)
                if length > max_len:
                    max_len = length
        adjusted = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def add_title_row(ws, title, subtitle=""):
    """タイトル行を追加"""
    ws.append([])  # 空行
    title_cell = ws.cell(row=ws.max_row, column=1, value=title)
    title_cell.font = Font(name="Arial", bold=True, size=16, color=COLORS["header_bg"])
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    
    if subtitle:
        ws.append([subtitle])
        sub_cell = ws.cell(row=ws.max_row, column=1)
        sub_cell.font = Font(name="Arial", size=10, color="666666", italic=True)
    
    ws.append([])  # 空行


def create_fhr_sheet(wb, fhr_data, currency="USD"):
    """Sheet 1: FHR料金一覧"""
    ws = wb.create_sheet("FHR料金一覧")
    ws.sheet_view.showGridLines = False
    
    # タイトル
    add_title_row(ws, "Fine Hotels & Resorts (FHR) 料金一覧",
                  f"※FHRは旅行期間中チェックアウト最大2回まで | 生成日: {datetime.now().strftime('%Y/%m/%d')}")
    
    # ヘッダー
    headers = [
        "ホテル名", "チェーン", "優先度", "エリア", "住所",
        "チェックイン", "チェックアウト", "泊数",
        f"1泊料金({currency})", f"合計料金({currency})",
        "特別オファー", f"実質1泊単価({currency})", "節約額", "備考"
    ]
    header_row = ws.max_row + 1
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        apply_header_style(cell, header)
    
    ws.row_dimensions[header_row].height = 40
    ws.freeze_panes = f"A{header_row + 1}"
    
    # データ行
    min_rate = min((h.get("effective_nightly_rate", 9999) for h in fhr_data), default=9999)
    
    for hotel in fhr_data:
        row = ws.max_row + 1
        chain_name = hotel.get("chain", "")
        has_offer = bool(hotel.get("special_offer", ""))
        is_cheapest = hotel.get("effective_nightly_rate", 9999) <= min_rate * 1.05
        
        bg = COLORS["offer_bg"] if has_offer else COLORS["fhr_bg"]
        chain_color = get_chain_color(chain_name)
        priority_str = "◎" if get_chain_priority(chain_name) <= 2 else \
                       "○" if get_chain_priority(chain_name) <= 4 else "△"
        
        savings = hotel.get("nightly_rate", 0) * hotel.get("nights", 1) - \
                  hotel.get("effective_total", hotel.get("total_rate", 0))
        
        row_data = [
            hotel.get("hotel_name", ""),
            chain_name,
            priority_str,
            hotel.get("area", ""),
            hotel.get("address", ""),
            hotel.get("check_in", ""),
            hotel.get("check_out", ""),
            hotel.get("nights", 1),
            hotel.get("nightly_rate", 0),
            hotel.get("total_rate", 0),
            hotel.get("special_offer", ""),
            hotel.get("effective_nightly_rate", hotel.get("nightly_rate", 0)),
            savings if savings > 0 else "",
            hotel.get("notes", ""),
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            num_fmt = None
            if col_idx in [9, 10, 12]:  # 料金列
                num_fmt = "#,##0"
                if is_cheapest and col_idx == 12:
                    cell.fill = PatternFill(start_color=COLORS["highlight"],
                                           end_color=COLORS["highlight"], fill_type="solid")
                    apply_data_style(cell, bg_color=COLORS["highlight"], bold=True,
                                   number_format=num_fmt, align="right")
                    continue
                apply_data_style(cell, bg_color=bg, number_format=num_fmt, align="right")
            elif col_idx == 13:  # 節約額
                num_fmt = '#,##0;-#,##0;""'
                apply_data_style(cell, bg_color=bg, font_color="006400" if savings > 0 else "000000",
                               number_format=num_fmt, align="right")
            elif col_idx == 2:  # チェーン名
                bold = get_chain_priority(chain_name) <= 4
                apply_data_style(cell, bg_color=bg, bold=bold, font_color=chain_color)
            elif col_idx == 3:  # 優先度
                apply_data_style(cell, bg_color=bg, align="center",
                               bold=(priority_str == "◎"))
            elif col_idx in [6, 7]:  # 日付
                apply_data_style(cell, bg_color=bg, align="center")
            elif col_idx == 8:  # 泊数
                apply_data_style(cell, bg_color=bg, align="center")
            else:
                apply_data_style(cell, bg_color=bg)
        
        ws.row_dimensions[row].height = 22
    
    # 凡例
    ws.append([])
    ws.append(["■ 凡例"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=10)
    legends = [
        ("背景色", COLORS["offer_bg"], "特別オファーあり"),
        ("背景色", COLORS["fhr_bg"], "通常FHR料金"),
        ("背景色", COLORS["highlight"], "最安値（±5%以内）"),
        ("優先度", None, "◎: Marriott/IHG  ○: Hyatt/Hilton  △: その他"),
    ]
    for leg_type, color, desc in legends:
        ws.append(["", leg_type, desc])
        if color:
            ws.cell(row=ws.max_row, column=3).fill = PatternFill(
                start_color=color, end_color=color, fill_type="solid")
    
    auto_fit_columns(ws)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{ws.max_row}"
    
    return ws


def create_thc_sheet(wb, thc_data, currency="USD"):
    """Sheet 2: THC料金一覧"""
    ws = wb.create_sheet("THC料金一覧")
    ws.sheet_view.showGridLines = False
    
    add_title_row(ws, "The Hotel Collection (THC) 料金一覧",
                  f"※THCは旅行期間中チェックアウト最大2回まで・各滞在2泊以上必須 | 生成日: {datetime.now().strftime('%Y/%m/%d')}")
    
    headers = [
        "ホテル名", "チェーン", "優先度", "エリア", "住所",
        "チェックイン", "チェックアウト", "泊数",
        f"1泊料金({currency})", f"合計料金({currency})",
        "特別オファー", f"実質1泊単価({currency})", "節約額", "備考"
    ]
    header_row = ws.max_row + 1
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        apply_header_style(cell, header)
    ws.row_dimensions[header_row].height = 40
    ws.freeze_panes = f"A{header_row + 1}"
    
    min_rate = min((h.get("effective_nightly_rate", 9999) for h in thc_data), default=9999)
    
    for hotel in thc_data:
        row = ws.max_row + 1
        chain_name = hotel.get("chain", "")
        has_offer = bool(hotel.get("special_offer", ""))
        is_cheapest = hotel.get("effective_nightly_rate", 9999) <= min_rate * 1.05
        
        bg = COLORS["offer_bg"] if has_offer else COLORS["thc_bg"]
        chain_color = get_chain_color(chain_name)
        priority_str = "◎" if get_chain_priority(chain_name) <= 2 else \
                       "○" if get_chain_priority(chain_name) <= 4 else "△"
        
        savings = hotel.get("nightly_rate", 0) * hotel.get("nights", 2) - \
                  hotel.get("effective_total", hotel.get("total_rate", 0))
        
        row_data = [
            hotel.get("hotel_name", ""),
            chain_name,
            priority_str,
            hotel.get("area", ""),
            hotel.get("address", ""),
            hotel.get("check_in", ""),
            hotel.get("check_out", ""),
            hotel.get("nights", 2),
            hotel.get("nightly_rate", 0),
            hotel.get("total_rate", 0),
            hotel.get("special_offer", ""),
            hotel.get("effective_nightly_rate", hotel.get("nightly_rate", 0)),
            savings if savings > 0 else "",
            hotel.get("notes", ""),
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            num_fmt = None
            if col_idx in [9, 10, 12]:
                num_fmt = "#,##0"
                if is_cheapest and col_idx == 12:
                    apply_data_style(cell, bg_color=COLORS["highlight"], bold=True,
                                   number_format=num_fmt, align="right")
                    continue
                apply_data_style(cell, bg_color=bg, number_format=num_fmt, align="right")
            elif col_idx == 13:
                num_fmt = '#,##0;-#,##0;""'
                apply_data_style(cell, bg_color=bg, font_color="006400" if savings > 0 else "000000",
                               number_format=num_fmt, align="right")
            elif col_idx == 2:
                bold = get_chain_priority(chain_name) <= 4
                apply_data_style(cell, bg_color=bg, bold=bold, font_color=chain_color)
            elif col_idx == 3:
                apply_data_style(cell, bg_color=bg, align="center",
                               bold=(priority_str == "◎"))
            elif col_idx in [6, 7]:
                apply_data_style(cell, bg_color=bg, align="center")
            elif col_idx == 8:
                apply_data_style(cell, bg_color=bg, align="center")
            else:
                apply_data_style(cell, bg_color=bg)
        
        ws.row_dimensions[row].height = 22
    
    auto_fit_columns(ws)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{ws.max_row}"
    
    return ws


def create_optimal_plan_sheet(wb, plan_data, total_info, currency="USD"):
    """Sheet 3: 最適プラン"""
    ws = wb.create_sheet("★ 最適プラン")
    ws.sheet_view.showGridLines = False
    
    # ====== サマリーセクション ======
    add_title_row(ws, "★ 推奨旅行プラン（最適化済み）",
                  "コスト最小・移動距離最小・会員資格特典最大化")
    
    # サマリー情報
    summary_items = [
        ("旅行総コスト", f"{currency} {total_info.get('total_cost', 0):,.0f}"),
        ("特典・オファー節約額", f"{currency} {total_info.get('total_savings', 0):,.0f}"),
        ("実質負担額（概算）", f"{currency} {total_info.get('net_cost', 0):,.0f}"),
        ("総泊数", f"{total_info.get('total_nights', 0)} 泊"),
        ("FHR滞在回数", f"{total_info.get('fhr_stays', 0)} 回（最大2回）"),
        ("THC滞在回数", f"{total_info.get('thc_stays', 0)} 回（最大2回）"),
        ("エリア移動回数", f"{total_info.get('area_changes', 0)} 回"),
    ]
    
    for label, value in summary_items:
        row = ws.max_row + 1
        label_cell = ws.cell(row=row, column=1, value=label)
        label_cell.font = Font(name="Arial", bold=True, size=11, color=COLORS["header_bg"])
        label_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        value_cell = ws.cell(row=row, column=2, value=value)
        value_cell.font = Font(name="Arial", size=11, bold=True)
        value_cell.alignment = Alignment(horizontal="left", vertical="center")
        
        if "コスト" in label or "節約" in label or "負担" in label:
            value_cell.font = Font(name="Arial", size=12, bold=True,
                                   color="006400" if "節約" in label else COLORS["header_bg"])
    
    ws.append([])
    ws.append(["─" * 60])  # 区切り線
    ws.append([])
    
    # ====== プラン詳細テーブル ======
    plan_title_row = ws.max_row + 1
    ws.cell(row=plan_title_row, column=1, value="【滞在スケジュール詳細】").font = \
        Font(name="Arial", bold=True, size=13, color=COLORS["header_bg"])
    ws.append([])
    
    headers = [
        "滞在順", "カテゴリ", "ホテル名", "チェーン",
        "チェックイン", "チェックアウト", "泊数",
        f"1泊単価({currency})", f"合計({currency})",
        "特別オファー", "エリア", "前拠点からの移動",
        "FHR付帯特典", "会員資格特典", "備考"
    ]
    
    header_row = ws.max_row + 1
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        apply_header_style(cell, header)
    ws.row_dimensions[header_row].height = 40
    ws.freeze_panes = f"A{header_row + 1}"
    
    # プランデータ
    stay_num = 1
    for stay in plan_data:
        row = ws.max_row + 1
        category = stay.get("category", "FHR")
        chain_name = stay.get("chain", "")
        bg = COLORS["fhr_bg"] if category == "FHR" else COLORS["thc_bg"]
        has_offer = bool(stay.get("special_offer", ""))
        if has_offer:
            bg = COLORS["offer_bg"]
        
        chain_color = get_chain_color(chain_name)
        
        row_data = [
            f"{stay_num}泊目〜",
            category,
            stay.get("hotel_name", ""),
            chain_name,
            stay.get("check_in", ""),
            stay.get("check_out", ""),
            stay.get("nights", 1),
            stay.get("effective_nightly_rate", stay.get("nightly_rate", 0)),
            stay.get("effective_total", stay.get("total_rate", 0)),
            stay.get("special_offer", "なし"),
            stay.get("area", ""),
            stay.get("travel_from_prev", "—"),
            stay.get("fhr_benefits", ""),
            stay.get("member_benefits", ""),
            stay.get("notes", ""),
        ]
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            if col_idx in [8, 9]:
                apply_data_style(cell, bg_color=bg, number_format="#,##0", align="right")
            elif col_idx == 2:  # カテゴリ
                cat_color = COLORS["marriott"] if category == "FHR" else COLORS["ihg"]
                apply_data_style(cell, bg_color=bg, bold=True, font_color=cat_color, align="center")
            elif col_idx == 4:  # チェーン
                apply_data_style(cell, bg_color=bg, bold=True, font_color=chain_color)
            elif col_idx in [5, 6]:  # 日付
                apply_data_style(cell, bg_color=bg, align="center")
            elif col_idx == 7:  # 泊数
                apply_data_style(cell, bg_color=bg, align="center")
            elif col_idx == 1:  # 滞在順
                apply_data_style(cell, bg_color=COLORS["plan_bg"], bold=True, align="center")
            elif col_idx == 12:  # 移動
                apply_data_style(cell, bg_color=bg, font_color="666666")
            else:
                apply_data_style(cell, bg_color=bg)
        
        ws.row_dimensions[row].height = 25
        stay_num += stay.get("nights", 1)
    
    # 合計行
    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=1, value="合計").font = Font(bold=True, size=11)
    ws.cell(row=total_row, column=7, value=total_info.get("total_nights", 0))
    ws.cell(row=total_row, column=9, value=total_info.get("total_cost", 0))
    for col_idx in [1, 7, 9]:
        cell = ws.cell(row=total_row, column=col_idx)
        cell.fill = PatternFill(start_color=COLORS["plan_bg"],
                               end_color=COLORS["plan_bg"], fill_type="solid")
        cell.font = Font(bold=True, size=11)
        cell.border = Border(top=Side(style="medium", color=COLORS["header_bg"]))
    ws.cell(row=total_row, column=9).number_format = "#,##0"
    
    # ====== 注意事項 ======
    ws.append([])
    ws.append([])
    notes_title = ws.cell(row=ws.max_row, column=1, value="【注意事項・ご確認ください】")
    notes_title.font = Font(bold=True, size=11, color="CC0000")
    
    notes = [
        "✓ FHRはアメックス トラベル経由で予約した場合のみ特典が適用されます",
        "✓ THCは2泊以上の滞在が必須条件です（1泊では利用不可）",
        "✓ Marriott/IHG/Hyatt/Hiltonポイントの付与可否は各ホテルに要確認",
        "✓ IHGアンバサダーのラウンジアクセスはラウンジのあるホテルのみ有効",
        "✓ 特別オファーの適用条件（予約期限・最小泊数等）は必ず確認してください",
        "✓ 料金は検索時点のものです。予約時に変動する場合があります",
        "✓ チェックイン/アウト時間はホテルにより異なります",
    ]
    for note in notes:
        ws.append([note])
        ws.cell(row=ws.max_row, column=1).font = Font(size=10, color="333333")
    
    auto_fit_columns(ws)
    
    return ws


def create_map_sheet(wb, hotel_addresses, area_groups):
    """Sheet 4: エリア・住所情報"""
    ws = wb.create_sheet("エリア情報")
    ws.sheet_view.showGridLines = False
    
    add_title_row(ws, "ホテル位置情報・エリアマップ",
                  "移動距離最小化のためのエリア分析")
    
    # エリアグループ別にホテルを表示
    for area_name, hotels in area_groups.items():
        if area_name.startswith("_"):  # 内部キー（_distances等）はスキップ
            continue
        row = ws.max_row + 1
        area_cell = ws.cell(row=row, column=1, value=f"■ {area_name}")
        area_cell.font = Font(name="Arial", bold=True, size=12, color=COLORS["header_bg"])
        area_cell.fill = PatternFill(start_color=COLORS["light_gray"],
                                     end_color=COLORS["light_gray"], fill_type="solid")
        
        # ヘッダー
        sub_headers = ["ホテル名", "カテゴリ", "チェーン", "住所", "最寄り交通"]
        header_row = ws.max_row + 1
        for col_idx, h in enumerate(sub_headers, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.value = h
            cell.font = Font(name="Arial", bold=True, size=10, color=COLORS["white"])
            cell.fill = PatternFill(start_color=COLORS["sub_header"],
                                   end_color=COLORS["sub_header"], fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        for hotel in hotels:
            row = ws.max_row + 1
            category = hotel.get("category", "")
            bg = COLORS["fhr_bg"] if category == "FHR" else COLORS["thc_bg"]
            
            hotel_row = [
                hotel.get("hotel_name", ""),
                category,
                hotel.get("chain", ""),
                hotel.get("address", ""),
                hotel.get("nearest_transit", ""),
            ]
            for col_idx, val in enumerate(hotel_row, 1):
                cell = ws.cell(row=row, column=col_idx, value=val)
                apply_data_style(cell, bg_color=bg)
        
        ws.append([])
    
    # 移動距離表
    ws.append([])
    ws.append(["【エリア間移動の目安】"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=11, color=COLORS["header_bg"])
    
    dist_headers = ["出発エリア", "到着エリア", "移動手段", "所要時間（目安）"]
    header_row = ws.max_row + 1
    for col_idx, h in enumerate(dist_headers, 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = h
        cell.font = Font(name="Arial", bold=True, size=10, color=COLORS["white"])
        cell.fill = PatternFill(start_color=COLORS["sub_header"],
                               end_color=COLORS["sub_header"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    distances = area_groups.get("_distances", []) if isinstance(area_groups, dict) else []
    for dist_item in distances:
        if isinstance(dist_item, (list, tuple)) and len(dist_item) == 4:
            from_area, to_area, transport, duration = dist_item
        else:
            continue
        row = ws.max_row + 1
        for col_idx, val in enumerate([from_area, to_area, transport, duration], 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            apply_data_style(cell, bg_color=COLORS["light_gray"])
    
    auto_fit_columns(ws)
    return ws


def generate_excel(data, output_path):
    """メイン関数: エクセルファイルを生成"""
    
    # データ検証
    fhr_data = data.get("fhr", [])
    thc_data = data.get("thc", [])
    plan_data = data.get("plan", [])
    total_info = data.get("total_info", {})
    hotel_addresses = data.get("hotel_addresses", {})
    area_groups = data.get("area_groups", {})
    currency = data.get("currency", "USD")
    
    # Workbook作成
    wb = openpyxl.Workbook()
    
    # デフォルトシートを削除
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # 各シートを作成
    if fhr_data:
        create_fhr_sheet(wb, fhr_data, currency)
        print(f"✓ FHR料金一覧シート作成完了 ({len(fhr_data)}件)")
    
    if thc_data:
        create_thc_sheet(wb, thc_data, currency)
        print(f"✓ THC料金一覧シート作成完了 ({len(thc_data)}件)")
    
    if plan_data:
        create_optimal_plan_sheet(wb, plan_data, total_info, currency)
        print(f"✓ 最適プランシート作成完了 ({len(plan_data)}滞在)")
    
    if hotel_addresses or area_groups:
        create_map_sheet(wb, hotel_addresses, area_groups)
        print(f"✓ エリア情報シート作成完了")
    
    # ファイル保存
    wb.save(output_path)
    print(f"\n✅ エクセルファイルを保存しました: {output_path}")
    return output_path


def create_sample_data():
    """サンプルデータを生成（テスト用）"""
    return {
        "currency": "USD",
        "fhr": [
            {
                "hotel_name": "Park Hyatt Tokyo",
                "chain": "Hyatt",
                "area": "西新宿",
                "address": "3-7-1-2 Nishi Shinjuku, Shinjuku, Tokyo",
                "check_in": "2025-03-05",
                "check_out": "2025-03-06",
                "nights": 1,
                "nightly_rate": 650,
                "total_rate": 650,
                "special_offer": "",
                "effective_nightly_rate": 650,
                "effective_total": 650,
                "notes": "",
            },
            {
                "hotel_name": "Aman Tokyo",
                "chain": "Aman",
                "area": "大手町",
                "address": "1-5-6 Otemachi, Chiyoda, Tokyo",
                "check_in": "2025-03-08",
                "check_out": "2025-03-09",
                "nights": 1,
                "nightly_rate": 1200,
                "total_rate": 1200,
                "special_offer": "3rd Night Free",
                "effective_nightly_rate": 800,
                "effective_total": 2400,
                "notes": "3泊目無料オファー適用",
            },
        ],
        "thc": [
            {
                "hotel_name": "InterContinental Tokyo Bay",
                "chain": "IHG",
                "area": "竹芝/浜松町",
                "address": "1-16-2 Kaigan, Minato, Tokyo",
                "check_in": "2025-03-01",
                "check_out": "2025-03-03",
                "nights": 2,
                "nightly_rate": 320,
                "total_rate": 640,
                "special_offer": "",
                "effective_nightly_rate": 320,
                "effective_total": 640,
                "notes": "IHGダイヤモンド・ラウンジアクセス",
            },
        ],
        "plan": [
            {
                "hotel_name": "InterContinental Tokyo Bay",
                "category": "THC",
                "chain": "IHG",
                "check_in": "2025-03-01",
                "check_out": "2025-03-03",
                "nights": 2,
                "nightly_rate": 320,
                "total_rate": 640,
                "special_offer": "",
                "effective_nightly_rate": 320,
                "effective_total": 640,
                "area": "竹芝/浜松町",
                "travel_from_prev": "—（旅行開始）",
                "fhr_benefits": "ホテルクレジット $100",
                "member_benefits": "IHGダイヤモンド：ラウンジ無料・スイートアップグレード申請",
                "notes": "",
            },
            {
                "hotel_name": "Park Hyatt Tokyo",
                "category": "FHR",
                "chain": "Hyatt",
                "check_in": "2025-03-05",
                "check_out": "2025-03-06",
                "nights": 1,
                "nightly_rate": 650,
                "total_rate": 650,
                "special_offer": "",
                "effective_nightly_rate": 650,
                "effective_total": 650,
                "area": "西新宿",
                "travel_from_prev": "竹芝→新宿（電車30分）",
                "fhr_benefits": "朝食2名・施設クレジット $100・アップグレード",
                "member_benefits": "Hyattグローバリスト：スイートアップグレード・午後4時チェックアウト",
                "notes": "",
            },
        ],
        "total_info": {
            "total_cost": 1290,
            "total_savings": 400,
            "net_cost": 890,
            "total_nights": 3,
            "fhr_stays": 1,
            "thc_stays": 1,
            "area_changes": 1,
        },
        "hotel_addresses": {},
        "area_groups": {
            "竹芝/浜松町": [
                {"hotel_name": "InterContinental Tokyo Bay", "category": "THC",
                 "chain": "IHG", "address": "1-16-2 Kaigan, Minato, Tokyo", "nearest_transit": "竹芝駅 徒歩5分"},
            ],
            "西新宿": [
                {"hotel_name": "Park Hyatt Tokyo", "category": "FHR",
                 "chain": "Hyatt", "address": "3-7-1-2 Nishi Shinjuku, Shinjuku", "nearest_transit": "西新宿駅 徒歩10分"},
            ],
            "_distances": [
                ("竹芝/浜松町", "西新宿", "電車（JR浜松町→新宿）", "約30分"),
            ],
        },
    }


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Amex Travel Planner - Excel Generator")
    parser.add_argument("--data", type=str, help="JSONデータファイルのパス")
    parser.add_argument("--json", type=str, help="JSONデータ文字列")
    parser.add_argument("--output", type=str, default="/mnt/user-data/outputs/amex_travel_plan.xlsx",
                       help="出力ファイルパス")
    parser.add_argument("--sample", action="store_true", help="サンプルデータでテスト実行")
    
    args = parser.parse_args()
    
    if args.sample:
        data = create_sample_data()
    elif args.data:
        with open(args.data, "r", encoding="utf-8") as f:
            data = json.load(f)
    elif args.json:
        data = json.loads(args.json)
    else:
        print("サンプルモードで実行します...")
        data = create_sample_data()
    
    output_path = Path(args.output)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    generate_excel(data, str(output_path))
